import os
import io
import re
import logging
import time
import json
from pathlib import Path
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import pdfplumber
import openpyxl
from src.embeddings import add_document, get_chroma_client

logger = logging.getLogger(__name__)

SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]
FOLDER_NAME = "normas para bot"
PROCESSED_FILE = Path("/tmp/processed_files.json")

SUPPORTED_MIMETYPES = {
    "application/pdf": "pdf",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xlsx",
    "application/vnd.ms-excel": "xlsx",
    "text/plain": "txt",
}


def detect_doc_type_and_date(filename: str) -> tuple:
    name = filename.upper()
    boletin_match = re.search(r'(\d{1,3})[/-](\d{4})', name)
    if "BOLETIN" in name or "INFORMACION" in name or boletin_match:
        if boletin_match:
            num = int(boletin_match.group(1))
            year = int(boletin_match.group(2))
            date_score = year * 1000 + num
        else:
            date_score = 0
        return "boletin", date_score
    norma_match = re.search(r'(ENERO|FEBRERO|MARZO|ABRIL|MAYO|JUNIO|JULIO|AGOSTO|SEPTIEMBRE|OCTUBRE|NOVIEMBRE|DICIEMBRE)[^0-9]*(\d{4})', name)
    if "NORMA" in name or norma_match:
        if norma_match:
            meses = {"ENERO":1,"FEBRERO":2,"MARZO":3,"ABRIL":4,"MAYO":5,"JUNIO":6,
                     "JULIO":7,"AGOSTO":8,"SEPTIEMBRE":9,"OCTUBRE":10,"NOVIEMBRE":11,"DICIEMBRE":12}
            mes = meses.get(norma_match.group(1), 0)
            year = int(norma_match.group(2))
            date_score = year * 100 + mes
        else:
            date_score = 0
        return "norma", date_score
    return "documento", 0


def get_drive_service():
    sa_env = os.environ["GOOGLE_SERVICE_ACCOUNT_JSON"]
    if sa_env.strip().startswith("{"):
        info = json.loads(sa_env)
        creds = Credentials.from_service_account_info(info, scopes=SCOPES)
    else:
        creds = Credentials.from_service_account_file(sa_env, scopes=SCOPES)
    return build("drive", "v3", credentials=creds)


def load_processed_files():
    if PROCESSED_FILE.exists():
        with open(PROCESSED_FILE) as f:
            return json.load(f)
    return {}


def save_processed_files(processed):
    PROCESSED_FILE.parent.mkdir(exist_ok=True)
    with open(PROCESSED_FILE, "w") as f:
        json.dump(processed, f)


def get_folder_id(service):
    result = service.files().list(
        q=f"name='{FOLDER_NAME}' and mimeType='application/vnd.google-apps.folder' and trashed=false",
        fields="files(id, name)"
    ).execute()
    files = result.get("files", [])
    if not files:
        raise ValueError(f"Carpeta '{FOLDER_NAME}' no encontrada")
    return files[0]["id"]


def list_files_in_folder(service, folder_id):
    mime_conditions = " or ".join([f"mimeType='{m}'" for m in SUPPORTED_MIMETYPES])
    result = service.files().list(
        q=f"'{folder_id}' in parents and ({mime_conditions}) and trashed=false",
        fields="files(id, name, modifiedTime, mimeType)"
    ).execute()
    return result.get("files", [])


def download_file(service, file_id):
    request = service.files().get_media(fileId=file_id)
    buffer = io.BytesIO()
    downloader = MediaIoBaseDownload(buffer, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    buffer.seek(0)
    return buffer


def extract_text_from_pdf(buffer):
    chunks = []
    with pdfplumber.open(buffer) as pdf:
        for i, page in enumerate(pdf.pages):
            text = page.extract_text()
            if text and text.strip():
                chunks.append({"page": i + 1, "text": text.strip(), "doc_type": None, "date_score": None})
    return chunks


def extract_text_from_excel(buffer):
    chunks = []
    wb = openpyxl.load_workbook(buffer, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_text = []
        for row in ws.iter_rows(values_only=True):
            row_values = [str(cell) for cell in row if cell is not None and str(cell).strip()]
            if row_values:
                rows_text.append(" | ".join(row_values))
        if rows_text:
            for i in range(0, len(rows_text), 50):
                chunk_text = "\n".join(rows_text[i:i+50])
                chunks.append({"page": f"{sheet_name}-{i//50 + 1}", "text": chunk_text, "doc_type": None, "date_score": None})
    wb.close()
    return chunks


def extract_text_from_txt(buffer):
    """Parser especial para el archivo de boletines COLFARMA."""
    text = buffer.read().decode("utf-8", errors="ignore")
    chunks = []

    # Dividir por separador de boletín
    sections = re.split(r'={40,}', text)

    for section in sections:
        section = section.strip()
        if len(section) < 50:
            continue

        # Extraer número y fecha del boletín
        num_match = re.search(r'INFORMACIÓN OBRAS SOCIALES\s+Nº\s+(\d+)/(\d{4})', section, re.IGNORECASE)
        fecha_match = re.search(r'Fecha:\s*(\d{4}-\d{2}-\d{2})', section)

        if num_match and fecha_match:
            num = int(num_match.group(1))
            year = int(num_match.group(2))
            date_score = year * 1000 + num
            doc_type = "boletin"
            page_id = f"boletin_{year}_{num:04d}"
        else:
            date_score = 0
            doc_type = "documento"
            page_id = f"txt_{len(chunks)}"

        chunks.append({
            "page": page_id,
            "text": section,
            "doc_type": doc_type,
            "date_score": date_score
        })

    return chunks


def process_new_files():
    logger.info("Revisando carpeta Drive por archivos nuevos...")
    service = get_drive_service()
    processed = load_processed_files()
    folder_id = get_folder_id(service)
    files = list_files_in_folder(service, folder_id)

    new_count = 0
    for f in files:
        file_id = f["id"]
        modified = f["modifiedTime"]
        mime = f["mimeType"]

        if processed.get(file_id) == modified:
            continue

        logger.info(f"Procesando: {f['name']}")
        try:
            buffer = download_file(service, file_id)
            default_type, default_score = detect_doc_type_and_date(f["name"])

            if mime == "application/pdf":
                chunks = extract_text_from_pdf(buffer)
            elif mime in ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/vnd.ms-excel"):
                chunks = extract_text_from_excel(buffer)
            else:
                chunks = extract_text_from_txt(buffer)

            for chunk in chunks:
                doc_id = f"{file_id}_p{chunk['page']}"
                doc_type = chunk.get("doc_type") or default_type
                date_score = chunk.get("date_score") if chunk.get("date_score") is not None else default_score
                metadata = {
                    "source": f["name"],
                    "file_id": file_id,
                    "page": str(chunk["page"]),
                    "doc_type": doc_type,
                    "date_score": str(date_score)
                }
                add_document(doc_id, chunk["text"], metadata)

            processed[file_id] = modified
            save_processed_files(processed)
            new_count += 1
            logger.info(f"✅ {f['name']} procesado ({len(chunks)} secciones)")

        except Exception as e:
            logger.error(f"Error procesando {f['name']}: {e}")

    logger.info(f"Revisión completa. {new_count} archivos nuevos procesados.")
    return new_count


def watch_drive(interval_seconds=300):
    logger.info(f"Watcher iniciado. Revisando cada {interval_seconds}s")
    while True:
        try:
            process_new_files()
        except Exception as e:
            logger.error(f"Error en watcher: {e}")
        time.sleep(interval_seconds)
